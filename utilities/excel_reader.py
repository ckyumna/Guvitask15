from pathlib import Path

from openpyxl.reader.excel import load_workbook


class ExcelReader:

    @staticmethod
    def get_multiple_test_data(sheet_name):

        root_path = Path(__file__).parent.parent

        file_path = root_path / "test_data" / "login_data.xlsx"

        wb = load_workbook(file_path)

        sheet = wb[sheet_name]

        header = []

        for cell in sheet[1]:
            header.append(cell.value)

        test_data = []

        for row in sheet.iter_rows(min_row=2, values_only=True):

            row_data = dict(zip(header, row))

            test_data.append(row_data)

        wb.close()

        return test_data

    @staticmethod
    def write_result(sheet_name, row_num, result):

        root_path = Path(__file__).parent.parent

        file_path = root_path / "test_data" / "login_data.xlsx"

        wb = load_workbook(file_path)

        sheet = wb[sheet_name]

        sheet.cell(row=row_num, column=7).value = result

        wb.save(file_path)

        wb.close()